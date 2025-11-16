"""Tests for native type storage in spreadsheet exports (XLSX, ODS)."""

import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from zipfile import ZipFile

from craftable import export_table
from craftable.styles import XlsxStyle, OdsStyle


class TestExportNativeTypes(unittest.TestCase):
    """Test that spreadsheet formats store native types appropriately."""

    def setUp(self):
        """Check for optional dependencies."""
        try:
            import openpyxl  # noqa: F401

            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False

        try:
            import odf.opendocument  # noqa: F401

            self.odf_available = True
        except ImportError:
            self.odf_available = False

    def test_xlsx_integer_storage(self):
        """Test XLSX stores integers as native numbers."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "integers.xlsx"

        rows = [[42], [100], [-5]]
        headers = ["Count"]

        export_table(rows, header_row=headers, style=XlsxStyle(), file=path)

        wb = load_workbook(path)
        ws = wb.active

        # Verify integers are stored as numbers, not text
        self.assertIsInstance(ws["A2"].value, int)
        self.assertEqual(ws["A2"].value, 42)
        self.assertIsInstance(ws["A3"].value, int)
        self.assertEqual(ws["A3"].value, 100)
        self.assertIsInstance(ws["A4"].value, int)
        self.assertEqual(ws["A4"].value, -5)

    def test_xlsx_float_storage(self):
        """Test XLSX stores floats as native numbers."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "floats.xlsx"

        rows = [[3.14], [2.718], [-0.5]]
        headers = ["Value"]

        export_table(rows, header_row=headers, style=XlsxStyle(), file=path)

        wb = load_workbook(path)
        ws = wb.active

        # Verify floats are stored as numbers
        self.assertIsInstance(ws["A2"].value, (int, float))
        self.assertAlmostEqual(ws["A2"].value, 3.14, places=2)
        self.assertIsInstance(ws["A3"].value, (int, float))
        self.assertAlmostEqual(ws["A3"].value, 2.718, places=3)

    def test_xlsx_boolean_storage(self):
        """Test XLSX stores booleans as native bool type."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "booleans.xlsx"

        rows = [[True], [False], [True]]
        headers = ["Flag"]

        export_table(rows, header_row=headers, style=XlsxStyle(), file=path)

        wb = load_workbook(path)
        ws = wb.active

        # Verify booleans are stored as bool type
        self.assertIsInstance(ws["A2"].value, bool)
        self.assertEqual(ws["A2"].value, True)
        self.assertIsInstance(ws["A3"].value, bool)
        self.assertEqual(ws["A3"].value, False)

    def test_xlsx_date_storage(self):
        """Test XLSX stores dates as native datetime objects."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "dates.xlsx"

        test_date = date(2024, 12, 25)
        rows = [[test_date], [date(2025, 1, 1)]]
        headers = ["Date"]

        export_table(rows, header_row=headers, style=XlsxStyle(), file=path)

        wb = load_workbook(path)
        ws = wb.active

        # Verify dates are stored as datetime objects (Excel converts date to datetime)
        self.assertIsInstance(ws["A2"].value, (date, datetime))
        # Check the date part matches (Excel stores as datetime with time=00:00:00)
        if isinstance(ws["A2"].value, datetime):
            self.assertEqual(ws["A2"].value.date(), test_date)
        else:
            self.assertEqual(ws["A2"].value, test_date)

    def test_xlsx_datetime_storage(self):
        """Test XLSX stores datetimes as native datetime objects."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "datetimes.xlsx"

        test_dt = datetime(2024, 12, 25, 14, 30, 0)
        rows = [[test_dt]]
        headers = ["DateTime"]

        export_table(rows, header_row=headers, style=XlsxStyle(), file=path)

        wb = load_workbook(path)
        ws = wb.active

        # Verify datetimes are stored as datetime objects
        self.assertIsInstance(ws["A2"].value, datetime)
        self.assertEqual(ws["A2"].value, test_dt)

    def test_xlsx_percent_format_with_precision(self):
        """Test XLSX applies percent format with proper precision."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "percents.xlsx"

        rows = [[0.75], [0.123], [1.0]]
        headers = ["Percent"]
        col_defs = [".2%"]  # Two decimal places, percent format

        export_table(
            rows, header_row=headers, col_defs=col_defs, style=XlsxStyle(), file=path
        )

        wb = load_workbook(path)
        ws = wb.active

        # Verify values are numeric (not text)
        self.assertIsInstance(ws["A2"].value, (int, float))
        self.assertAlmostEqual(ws["A2"].value, 0.75, places=2)

        # Verify percent format is applied
        self.assertIn("%", ws["A2"].number_format)
        # Should have two decimal places
        self.assertIn("0.00%", ws["A2"].number_format)

    def test_xlsx_decimal_precision_format(self):
        """Test XLSX applies decimal format with specified precision."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "decimals.xlsx"

        rows = [[123.456], [789.012]]
        headers = ["Value"]
        col_defs = [".3f"]  # Three decimal places

        export_table(
            rows, header_row=headers, col_defs=col_defs, style=XlsxStyle(), file=path
        )

        wb = load_workbook(path)
        ws = wb.active

        # Verify values are numeric
        self.assertIsInstance(ws["A2"].value, (int, float))
        self.assertAlmostEqual(ws["A2"].value, 123.456, places=3)

        # Verify decimal format with 3 places
        self.assertIn("0.000", ws["A2"].number_format)

    def test_xlsx_grouping_separator_format(self):
        """Test XLSX applies thousand separator formatting."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "grouped.xlsx"

        rows = [[1234567], [9876543]]
        headers = ["Large Number"]
        col_defs = [",.0f"]  # Grouping with no decimals

        export_table(
            rows, header_row=headers, col_defs=col_defs, style=XlsxStyle(), file=path
        )

        wb = load_workbook(path)
        ws = wb.active

        # Verify values are numeric
        self.assertIsInstance(ws["A2"].value, (int, float))
        self.assertEqual(ws["A2"].value, 1234567)

        # Verify grouping format is applied
        self.assertIn("#,##0", ws["A2"].number_format)

    def test_xlsx_prefix_forces_text(self):
        """Test XLSX stores as text when prefix is present."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "with_prefix.xlsx"

        rows = [[42], [100]]
        headers = ["Count"]
        col_defs = ["$(10)"]  # Prefix $ forces text mode

        export_table(
            rows, header_row=headers, col_defs=col_defs, style=XlsxStyle(), file=path
        )

        wb = load_workbook(path)
        ws = wb.active

        # Verify stored as text because of prefix
        self.assertIsInstance(ws["A2"].value, str)
        self.assertIn("$", ws["A2"].value)
        self.assertIn("42", ws["A2"].value)

    def test_ods_integer_storage(self):
        """Test ODS stores integers with float value-type."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "integers.ods"

        rows = [[42], [100], [-5]]
        headers = ["Count"]

        export_table(rows, header_row=headers, style=OdsStyle(), file=path)

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should contain float value-type for integers
            self.assertIn('office:value-type="float"', content_xml)
            self.assertIn('office:value="42"', content_xml)
            self.assertIn('office:value="100"', content_xml)
            self.assertIn('office:value="-5"', content_xml)

    def test_ods_float_storage(self):
        """Test ODS stores floats with float value-type."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "floats.ods"

        rows = [[3.14], [2.718]]
        headers = ["Value"]

        export_table(rows, header_row=headers, style=OdsStyle(), file=path)

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should contain float value-type
            self.assertIn('office:value-type="float"', content_xml)
            self.assertIn('office:value="3.14"', content_xml)

    def test_ods_boolean_storage(self):
        """Test ODS stores booleans with boolean value-type."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "booleans.ods"

        rows = [[True], [False]]
        headers = ["Flag"]

        export_table(rows, header_row=headers, style=OdsStyle(), file=path)

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should contain boolean value-type
            self.assertIn('office:value-type="boolean"', content_xml)
            self.assertIn('office:value="true"', content_xml)
            self.assertIn('office:value="false"', content_xml)

    def test_ods_date_storage(self):
        """Test ODS stores dates with date value-type."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "dates.ods"

        test_date = date(2024, 12, 25)
        rows = [[test_date]]
        headers = ["Date"]

        export_table(rows, header_row=headers, style=OdsStyle(), file=path)

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should contain date value-type with ISO format
            self.assertIn('office:value-type="date"', content_xml)
            self.assertIn('office:date-value="2024-12-25"', content_xml)

    def test_ods_datetime_storage(self):
        """Test ODS stores datetimes with date value-type."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "datetimes.ods"

        test_dt = datetime(2024, 12, 25, 14, 30, 0)
        rows = [[test_dt]]
        headers = ["DateTime"]

        export_table(rows, header_row=headers, style=OdsStyle(), file=path)

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should contain date value-type with ISO format including time
            self.assertIn('office:value-type="date"', content_xml)
            self.assertIn("2024-12-25T14:30:00", content_xml)

    def test_ods_percent_storage(self):
        """Test ODS stores percentages with percentage value-type."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "percents.ods"

        rows = [[0.75], [0.123]]
        headers = ["Percent"]
        col_defs = [".2%"]  # Percent format

        export_table(
            rows, header_row=headers, col_defs=col_defs, style=OdsStyle(), file=path
        )

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should contain percentage value-type
            self.assertIn('office:value-type="percentage"', content_xml)
            self.assertIn('office:value="0.75"', content_xml)

    def test_ods_prefix_forces_text(self):
        """Test ODS stores as text when prefix is present."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        tmp_dir = tempfile.mkdtemp()
        path = Path(tmp_dir) / "with_prefix.ods"

        rows = [[42], [100]]
        headers = ["Count"]
        col_defs = ["$(10)"]  # Prefix $ forces text mode

        export_table(
            rows, header_row=headers, col_defs=col_defs, style=OdsStyle(), file=path
        )

        with ZipFile(path, "r") as zip_file:
            content_xml = zip_file.read("content.xml").decode("utf-8")

            # Should NOT contain value-type (text is default)
            # Should contain the prefix in text
            self.assertIn("$", content_xml)
            # Values with prefix should be text, not have office:value
            lines_with_prefix = [
                line
                for line in content_xml.split(">")
                if "$42" in line or "$100" in line
            ]
            self.assertTrue(len(lines_with_prefix) > 0)


if __name__ == "__main__":
    unittest.main()

"""Tests for optional binary export formats (DOCX, XLSX, ODT, ODS, RTF)."""

import io
import tempfile
import unittest
from pathlib import Path

from craftable import export_table
from craftable.styles import DocxStyle, XlsxStyle, OdtStyle, OdsStyle, RtfStyle


class TestExportOptionalFormats(unittest.TestCase):
    """Test optional export formats with backward compatibility."""

    ROWS = [["Alice", 30, 95000], ["Bob", 25, 75000]]
    HEADERS = ["Name", "Age", "Salary"]

    def setUp(self):
        """Set up temporary directory and check for optional dependencies."""
        self.tmp_dir = tempfile.mkdtemp()

        # Check optional dependencies
        try:
            import docx  # noqa: F401

            self.docx_available = True
        except ImportError:
            self.docx_available = False

        try:
            import openpyxl  # noqa: F401

            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False

        try:
            import odf.opendocument  # noqa: F401

            self.odf_available = True
        except ImportError:
            self.odf_available = False

    def test_rtf_basic(self):
        """Test basic RTF export to file."""
        path = Path(self.tmp_dir) / "table.rtf"
        export_table(self.ROWS, header_row=self.HEADERS, style=RtfStyle(), file=path)
        txt = path.read_text(encoding="utf-8")
        self.assertIn("{\\rtf1", txt)
        self.assertIn("\\trowd", txt)

    def test_docx_basic(self):
        """Test basic DOCX export to file."""
        if not self.docx_available:
            self.skipTest("python-docx not available")

        path = Path(self.tmp_dir) / "table.docx"
        export_table(self.ROWS, header_row=self.HEADERS, style=DocxStyle(), file=path)
        self.assertGreater(path.stat().st_size, 0)

    def test_xlsx_basic(self):
        """Test basic XLSX export to file."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        path = Path(self.tmp_dir) / "table.xlsx"
        export_table(self.ROWS, header_row=self.HEADERS, style=XlsxStyle(), file=path)
        self.assertGreater(path.stat().st_size, 0)

    def test_odt_basic(self):
        """Test basic ODT export to file."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        path = Path(self.tmp_dir) / "table.odt"
        export_table(self.ROWS, header_row=self.HEADERS, style=OdtStyle(), file=path)
        self.assertGreater(path.stat().st_size, 0)

    def test_ods_basic(self):
        """Test basic ODS export to file."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        path = Path(self.tmp_dir) / "table.ods"
        export_table(self.ROWS, header_row=self.HEADERS, style=OdsStyle(), file=path)
        self.assertGreater(path.stat().st_size, 0)

    def test_xlsx_sheet_name(self):
        """Test XLSX export with custom sheet name."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        from openpyxl import load_workbook

        path = Path(self.tmp_dir) / "custom.xlsx"
        export_table(
            self.ROWS,
            header_row=self.HEADERS,
            style=XlsxStyle(sheet_name="Custom"),
            file=path,
        )
        wb = load_workbook(path)
        self.assertIn("Custom", wb.sheetnames)

    def test_xlsx_stream_signature(self):
        """Test XLSX export to stream with PK signature check."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        stream = io.BytesIO()
        export_table(self.ROWS, header_row=self.HEADERS, style=XlsxStyle(), file=stream)
        self.assertGreater(stream.tell(), 0)
        stream.seek(0)
        self.assertEqual(stream.read(2), b"PK")

    def test_rtf_stream(self):
        """Test RTF export to stream."""
        stream = io.StringIO()
        export_table(self.ROWS, header_row=self.HEADERS, style=RtfStyle(), file=stream)
        txt = stream.getvalue()
        self.assertIn("{\\rtf1", txt)
        self.assertIn("\\trowd", txt)

    def test_docx_stream(self):
        """Test DOCX export to stream."""
        if not self.docx_available:
            self.skipTest("python-docx not available")

        stream = io.BytesIO()
        export_table(self.ROWS, header_row=self.HEADERS, style=DocxStyle(), file=stream)
        self.assertGreater(stream.tell(), 0)
        stream.seek(0)
        self.assertEqual(stream.read(2), b"PK")

    def test_odt_stream(self):
        """Test ODT export to stream."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        stream = io.BytesIO()
        export_table(self.ROWS, header_row=self.HEADERS, style=OdtStyle(), file=stream)
        self.assertGreater(stream.tell(), 0)
        stream.seek(0)
        self.assertEqual(stream.read(2), b"PK")

    def test_ods_stream(self):
        """Test ODS export to stream."""
        if not self.odf_available:
            self.skipTest("odfpy not available")

        stream = io.BytesIO()
        export_table(self.ROWS, header_row=self.HEADERS, style=OdsStyle(), file=stream)
        self.assertGreater(stream.tell(), 0)
        stream.seek(0)
        self.assertEqual(stream.read(2), b"PK")

    def test_xlsx_pre_post_processors(self):
        """Test XLSX with preprocessors and postprocessors."""
        if not self.openpyxl_available:
            self.skipTest("openpyxl not available")

        path = Path(self.tmp_dir) / "proc.xlsx"

        def double(v, row, idx):
            return v * 2 if isinstance(v, (int, float)) else v

        def exclaim(original, text, row, idx):
            return f"{text}!" if isinstance(original, str) else text

        export_table(
            self.ROWS,
            header_row=self.HEADERS,
            style=XlsxStyle(),
            file=path,
            preprocessors=[None, None, double],
            postprocessors=[exclaim, None, None],
        )
        self.assertGreater(path.stat().st_size, 0)

    def test_rtf_special_chars(self):
        """Test RTF export with special characters that need escaping."""
        path = Path(self.tmp_dir) / "special.rtf"
        data = [["Test\\Backslash", "Test{Brace}", "Test}End"]]
        export_table(data, header_row=["A", "B", "C"], style=RtfStyle(), file=path)
        txt = path.read_text(encoding="utf-8")
        # Expect escaped RTF sequences: backslashes doubled, braces escaped
        self.assertIn("Test\\\\Backslash", txt)
        self.assertIn("Test\\{Brace\\}", txt)
        self.assertIn("Test\\}End", txt)


if __name__ == "__main__":
    unittest.main()
